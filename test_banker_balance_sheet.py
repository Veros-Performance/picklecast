#!/usr/bin/env python3
"""
Unit tests for the banker-format Balance Sheet export (Y1 monthly + Y2 EOY).
Tests enforce correctness of structure and totals without being fragile about formatting.
"""

import pytest
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime
import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def create_sample_bs_data():
    """Create sample balance sheet data for testing"""
    # Y1 monthly data (12 months)
    bs_y1_monthly = []
    for month in range(1, 13):
        cash = 200000 + month * 10000
        ppe_net = 1214000 - 10000 * month
        total_assets = cash + ppe_net
        debt = 900000 - month * 5000
        equity = total_assets - debt  # Ensure balance sheet balances

        bs_y1_monthly.append({
            "month": month,
            "cash": cash,
            "ppe_gross": 1214000,
            "accumulated_depreciation": -10000 * month,
            "ppe_net": ppe_net,
            "total_assets": total_assets,
            "debt_balance": debt,
            "equity": equity,
            "total_liabilities_equity": total_assets,  # Should equal total_assets
            "check": 0.0
        })

    # Y2 EOY data
    bs_y2_eoy = {
        "month": 24,
        "cash": 450000,
        "ppe_gross": 1214000,
        "accumulated_depreciation": -240000,
        "ppe_net": 974000,
        "total_assets": 1424000,
        "debt_balance": 750000,
        "equity": 674000,
        "total_liabilities_equity": 1424000,
        "check": 0.0
    }

    return bs_y1_monthly, bs_y2_eoy


def get_balance_sheet_dataframe(bs_bytes):
    """Convert Balance Sheet bytes to DataFrame for testing"""
    workbook = openpyxl.load_workbook(BytesIO(bs_bytes))
    worksheet = workbook.active

    # Extract all data into a list
    data = []
    for row in worksheet.iter_rows(values_only=True):
        data.append(list(row))

    return pd.DataFrame(data)


def find_row_by_label(df, label):
    """Find row index by searching for label in column A"""
    for idx, row in df.iterrows():
        if pd.notna(row[0]) and label in str(row[0]).strip():
            return idx
    return None


class TestBalanceSheetStructure:
    """Test the structure and headers of the Balance Sheet"""

    def test_columns_and_headers(self):
        """Verify the export includes correct monthly columns and headers"""
        from app import create_banker_balance_sheet

        bs_y1, bs_y2 = create_sample_bs_data()
        bs_bytes = create_banker_balance_sheet(bs_y1, bs_y2)
        df = get_balance_sheet_dataframe(bs_bytes)

        current_year = datetime.now().year

        # Check Y1 monthly headers (row 1, columns C-N)
        month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                       'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

        for i, month_name in enumerate(month_names):
            expected_header = f"{month_name} {current_year}"
            actual_header = df.iloc[1, i + 2]  # Row 1, starting at column C
            assert actual_header == expected_header, f"Month {i+1} header mismatch: expected '{expected_header}', got '{actual_header}'"

        # Check Y2 EOY header (row 1, column O)
        y2_header = df.iloc[1, 14]  # Column O
        expected_y2 = f"EOY {current_year + 1}"
        assert y2_header == expected_y2, f"Y2 EOY header mismatch: expected '{expected_y2}', got '{y2_header}'"

    def test_date_rows(self):
        """Verify each header has a corresponding 'as of' date row"""
        from app import create_banker_balance_sheet

        bs_y1, bs_y2 = create_sample_bs_data()
        bs_bytes = create_banker_balance_sheet(bs_y1, bs_y2)
        df = get_balance_sheet_dataframe(bs_bytes)

        # Check that row 2 contains "as of" dates
        for col in range(2, 15):  # Columns C through O
            date_cell = df.iloc[2, col]
            assert pd.notna(date_cell), f"Date cell at column {col} is empty"
            assert "as of" in str(date_cell), f"Date format issue in column {col}: {date_cell}"

    def test_title(self):
        """Verify the title cell equals 'Balance Sheet (Projected)'"""
        from app import create_banker_balance_sheet

        bs_y1, bs_y2 = create_sample_bs_data()
        bs_bytes = create_banker_balance_sheet(bs_y1, bs_y2)
        df = get_balance_sheet_dataframe(bs_bytes)

        # Title should be in the first row
        title_found = False
        for col in range(df.shape[1]):
            cell = df.iloc[0, col]
            if pd.notna(cell) and "Balance Sheet (Projected)" in str(cell):
                title_found = True
                break

        assert title_found, "Title 'Balance Sheet (Projected)' not found in first row"


class TestBalanceSheetLabels:
    """Test that all required labels are present in correct order"""

    def test_required_labels_present(self):
        """Assert that all required section/total labels are present"""
        from app import create_banker_balance_sheet

        bs_y1, bs_y2 = create_sample_bs_data()
        bs_bytes = create_banker_balance_sheet(bs_y1, bs_y2)
        df = get_balance_sheet_dataframe(bs_bytes)

        required_labels = [
            "Assets",
            "Current Assets",
            "Cash",
            "Other current assets",
            "Total Current Assets",
            "Fixed Assets",
            "Fixed assets (gross)",
            "Other fixed assets",
            "(LESS accumulated depreciation",
            "Total Fixed Assets (net",
            "Other Assets",
            "Total Other Assets",
            "TOTAL Assets",
            "Liabilities and Equity",
            "Current Liabilities",
            "Payroll liabilities",
            "Other current liabilities",
            "Total Current Liabilities",
            "Total Long-term Debt",
            "Total Liabilities",
            "Owners' Equity",
            "Total Owners' Equity",
            "Total Liabilities & Equity"
        ]

        # Get all labels from column A
        all_labels = []
        for idx in range(df.shape[0]):
            cell = df.iloc[idx, 0]
            if pd.notna(cell):
                all_labels.append(str(cell).strip())

        # Check each required label
        for label in required_labels:
            found = any(label in cell for cell in all_labels)
            assert found, f"Required label '{label}' not found in Balance Sheet"

    def test_label_order(self):
        """Verify labels appear in the correct order"""
        from app import create_banker_balance_sheet

        bs_y1, bs_y2 = create_sample_bs_data()
        bs_bytes = create_banker_balance_sheet(bs_y1, bs_y2)
        df = get_balance_sheet_dataframe(bs_bytes)

        # Key labels that must appear in this order
        ordered_labels = [
            "Assets",
            "Current Assets",
            "Total Current Assets",
            "Fixed Assets",
            "Total Fixed Assets",
            "TOTAL Assets",
            "Liabilities and Equity",
            "Total Liabilities",
            "Total Liabilities & Equity"
        ]

        # Find positions
        positions = []
        for label in ordered_labels:
            row = find_row_by_label(df, label)
            assert row is not None, f"Label '{label}' not found"
            positions.append(row)

        # Verify order
        for i in range(len(positions) - 1):
            assert positions[i] < positions[i+1], f"Label order incorrect: '{ordered_labels[i]}' should come before '{ordered_labels[i+1]}'"


class TestBalanceSheetMath:
    """Test mathematical relationships and totals"""

    def test_total_current_assets(self):
        """Total Current Assets == Cash + Other current assets"""
        from app import create_banker_balance_sheet

        bs_y1, bs_y2 = create_sample_bs_data()
        bs_bytes = create_banker_balance_sheet(bs_y1, bs_y2)
        df = get_balance_sheet_dataframe(bs_bytes)

        cash_row = find_row_by_label(df, "Cash")
        other_current_row = find_row_by_label(df, "Other current assets")
        total_current_row = find_row_by_label(df, "Total Current Assets")

        assert all(row is not None for row in [cash_row, other_current_row, total_current_row])

        # Test for each month column (C through N) and Y2 EOY (O)
        for col in range(2, 15):
            cash = df.iloc[cash_row, col] or 0
            other = df.iloc[other_current_row, col] or 0
            total = df.iloc[total_current_row, col] or 0

            # Convert to float if numeric
            if isinstance(cash, (int, float)) and isinstance(total, (int, float)):
                expected = float(cash) + float(other)
                actual = float(total)
                assert abs(expected - actual) < 0.01, f"Column {col}: Total Current Assets mismatch"

    def test_total_fixed_assets(self):
        """Total Fixed Assets (net) == Fixed assets (gross) + Accumulated depreciation"""
        from app import create_banker_balance_sheet

        bs_y1, bs_y2 = create_sample_bs_data()
        bs_bytes = create_banker_balance_sheet(bs_y1, bs_y2)
        df = get_balance_sheet_dataframe(bs_bytes)

        gross_row = find_row_by_label(df, "Fixed assets (gross)")
        depreciation_row = find_row_by_label(df, "accumulated depreciation")
        net_row = find_row_by_label(df, "Total Fixed Assets (net")

        assert all(row is not None for row in [gross_row, depreciation_row, net_row])

        # Test for each month column
        for col in range(2, 15):
            gross = df.iloc[gross_row, col] or 0
            depreciation = df.iloc[depreciation_row, col] or 0
            net = df.iloc[net_row, col] or 0

            if isinstance(gross, (int, float)) and isinstance(net, (int, float)):
                expected = float(gross) + float(depreciation)  # depreciation is negative
                actual = float(net)
                assert abs(expected - actual) < 0.01, f"Column {col}: Total Fixed Assets (net) mismatch"

    def test_balance_sheet_equation(self):
        """TOTAL Assets == Total Liabilities & Equity (within tolerance)"""
        from app import create_banker_balance_sheet

        bs_y1, bs_y2 = create_sample_bs_data()
        bs_bytes = create_banker_balance_sheet(bs_y1, bs_y2)
        df = get_balance_sheet_dataframe(bs_bytes)

        assets_row = find_row_by_label(df, "TOTAL Assets")
        liab_equity_row = find_row_by_label(df, "Total Liabilities & Equity")

        assert assets_row is not None, "TOTAL Assets row not found"
        assert liab_equity_row is not None, "Total Liabilities & Equity row not found"

        # Test for each month column
        for col in range(2, 15):
            assets = df.iloc[assets_row, col]
            liab_equity = df.iloc[liab_equity_row, col]

            if isinstance(assets, (int, float)) and isinstance(liab_equity, (int, float)):
                diff = abs(float(assets) - float(liab_equity))
                # Allow small differences due to rounding in calculations
                assert diff < 0.01, f"Column {col}: Balance sheet doesn't balance (diff={diff})"


class TestBalanceSheetDefaults:
    """Test default values for optional line items"""

    def test_default_values(self):
        """Verify optional items exist with value 0 if not provided"""
        from app import create_banker_balance_sheet

        bs_y1, bs_y2 = create_sample_bs_data()
        bs_bytes = create_banker_balance_sheet(bs_y1, bs_y2)
        df = get_balance_sheet_dataframe(bs_bytes)

        # Items that should default to 0
        default_items = [
            "Other current assets",
            "Other fixed assets",
            "Total Other Assets",
            "Payroll liabilities",
            "Other current liabilities"
        ]

        for item in default_items:
            row = find_row_by_label(df, item)
            assert row is not None, f"Default item '{item}' not found"

            # Check that values are 0 for at least one column
            for col in range(2, 15):
                val = df.iloc[row, col]
                if isinstance(val, (int, float)):
                    assert val == 0, f"Default item '{item}' should be 0, got {val}"
                    break


class TestBalanceSheetIntegrity:
    """Test general integrity of the Balance Sheet"""

    def test_numeric_cells_are_numbers(self):
        """Verify all numeric cells are numbers (not strings)"""
        from app import create_banker_balance_sheet

        bs_y1, bs_y2 = create_sample_bs_data()
        bs_bytes = create_banker_balance_sheet(bs_y1, bs_y2)

        # Load with openpyxl to check cell types
        workbook = openpyxl.load_workbook(BytesIO(bs_bytes))
        worksheet = workbook.active

        # Check cells that should contain numbers (columns C-O, rows with data)
        for row in range(4, 30):  # Data rows approximately
            for col in range(3, 16):  # Columns C-O (1-indexed in openpyxl)
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None and cell.value != "":
                    # Skip date rows and header rows
                    if not isinstance(cell.value, str) or not ("as of" in str(cell.value)):
                        # If it looks numeric, it should be a number type
                        if str(cell.value).replace(',', '').replace('(', '').replace(')', '').replace('-', '').replace('.', '').isdigit():
                            assert isinstance(cell.value, (int, float)), f"Cell at row {row}, col {col} should be numeric but is {type(cell.value)}: {cell.value}"

    def test_no_missing_required_data(self):
        """Verify that key data points are not missing"""
        from app import create_banker_balance_sheet

        bs_y1, bs_y2 = create_sample_bs_data()
        bs_bytes = create_banker_balance_sheet(bs_y1, bs_y2)
        df = get_balance_sheet_dataframe(bs_bytes)

        # Key items that must have values
        required_items = [
            "Cash",
            "TOTAL Assets",
            "Total Long-term Debt",
            "Total Owners' Equity",
            "Total Liabilities & Equity"
        ]

        for item in required_items:
            row = find_row_by_label(df, item)
            assert row is not None, f"Required item '{item}' not found"

            # Check that at least Y1M1 and Y2 EOY have values
            y1m1_val = df.iloc[row, 2]  # Column C
            y2_eoy_val = df.iloc[row, 14]  # Column O

            assert pd.notna(y1m1_val), f"{item} missing value for Y1M1"
            assert pd.notna(y2_eoy_val), f"{item} missing value for Y2 EOY"


if __name__ == "__main__":
    # Run all tests
    pytest.main([__file__, "-v"])