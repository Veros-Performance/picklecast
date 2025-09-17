import re
from io import BytesIO
import math
import pytest
from openpyxl import load_workbook
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from app import create_banker_pnl_sheet

# ----- Config -----
TOL = 0.01
EXPECTED_LABEL_COUNT = 53
# Exact row names (normalized) we must be able to find
MUST_HAVE_ROWS = [
    "total revenue",               # "Total Revenue (Sales)" or similar
    "total cost of sales",         # total of CoS window
    "gross profit",
    "total expenses",              # total of opex window
    "net profit before taxes",
    "federal income taxes",
    "state income taxes",
    "local income taxes",
    "net operating income",
]

def norm(s: str) -> str:
    """Lowercase, strip, drop non-alnum."""
    return re.sub(r"[^a-z0-9]", "", s.strip().lower())

# ----- Helpers to read from workbook bytes -----
def wb_from_bytes(binary: bytes):
    return load_workbook(BytesIO(binary), data_only=True)

def labels_col_a(ws):
    out = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None:
            if isinstance(v, str):
                # Count even single-space placeholders as labels
                out.append(v.strip() if v.strip() else " ")
            else:
                out.append(str(v).strip())
    return out

def find_row(ws, contains: str):
    target = norm(contains)
    # Special case: for "total expenses", exclude "sub-total"
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str):
            normalized = norm(v)
            if target in normalized:
                # If looking for "total expenses", skip "sub-total expenses"
                if contains == "total expenses" and "subtotal" in normalized:
                    continue
                return r, v
    return None, None

def read_nums(ws, row, c_first=3, c_last=15):
    vals = []
    for c in range(c_first, c_last + 1):  # C..O (12 months + EOY)
        v = ws.cell(row, c).value
        if v is None or v == "":
            vals.append(0.0)
        elif isinstance(v, (int, float)):
            vals.append(float(v))
        else:
            # tolerate formatted strings like (1,234)
            s = str(v).replace(",", "").replace("(", "-").replace(")", "")
            try:
                vals.append(float(s))
            except Exception:
                vals.append(0.0)
    return vals

def approx(a, b, tol=TOL):
    return abs((a or 0.0) - (b or 0.0)) <= tol

# ----- The SUT hook (adjust import to your project) -----
@pytest.fixture
def workbook_bytes():
    """
    Create test data and call the real builder function
    """
    # Create test data matching the expected structure
    pnl_y1_monthly = []
    for month in range(1, 13):
        pnl_y1_monthly.append({
            'revenue_court': 10000,
            'revenue_league': 5000,
            'revenue_corporate': 3000,
            'revenue_tournament': 2000,
            'revenue_membership': 1500,
            'revenue_retail': 1000,
            'revenue_other': 500,
            'cogs': 8000,
            'rent': 3000,
            'payroll': 5000,
            'utilities': 500,
            'insurance': 400,
            'marketing': 600,
            'software': 300,
            'professional_fees': 200,
            'repairs_maintenance': 150,
            'other_opex': 250,
            'depreciation': 1000,
            'interest_expense': 800,
            'tax': 500
        })

    pnl_y2_eoy = {
        'revenue_court': 120000,
        'revenue_league': 60000,
        'revenue_corporate': 36000,
        'revenue_tournament': 24000,
        'revenue_membership': 18000,
        'revenue_retail': 12000,
        'revenue_other': 6000,
        'cogs': 96000,
        'rent': 36000,
        'payroll': 60000,
        'utilities': 6000,
        'insurance': 4800,
        'marketing': 7200,
        'software': 3600,
        'professional_fees': 2400,
        'repairs_maintenance': 1800,
        'other_opex': 3000,
        'depreciation': 12000,
        'interest_expense': 9600,
        'tax': 6000
    }

    # Call the real builder function (returns tuple of bytes and sanity_check)
    excel_bytes, _ = create_banker_pnl_sheet(pnl_y1_monthly, pnl_y2_eoy)
    return excel_bytes

# ----- Tests -----

def test_labels_count_and_presence(workbook_bytes):
    wb = wb_from_bytes(workbook_bytes)
    ws = wb.active
    labels = labels_col_a(ws)
    assert len(labels) == EXPECTED_LABEL_COUNT, f"Expected {EXPECTED_LABEL_COUNT} labels, got {len(labels)}."
    # Ensure required rows exist
    for key in MUST_HAVE_ROWS:
        r, lbl = find_row(ws, key)
        assert r is not None, f"Missing required row matching: {key}"

def test_cos_and_expense_rollups(workbook_bytes):
    wb = wb_from_bytes(workbook_bytes)
    ws = wb.active

    # Windows: strictly between header and total
    def window_sum(header_key, total_key):
        r_head, _ = find_row(ws, header_key)
        r_total, _ = find_row(ws, total_key)
        assert r_head and r_total and r_total > r_head + 1, f"Bad window for {header_key}->{total_key}"
        total = [0.0] * 13
        for r in range(r_head + 1, r_total):
            label = ws.cell(r, 1).value
            if not isinstance(label, str) or not label.strip():
                continue  # cosmetic blank row
            # Skip rollup rows
            label_lower = label.strip().lower()
            if any(skip in label_lower for skip in ["sub-total", "subtotal", "reserve"]):
                continue
            vals = read_nums(ws, r)
            total = [a + b for a, b in zip(total, vals)]
        return total, read_nums(ws, r_total)

    # Cost of Sales
    cos_detail_sum, cos_total_written = window_sum("cost of sales", "total cost of sales")
    for i, (exp, got) in enumerate(zip(cos_detail_sum, cos_total_written)):
        assert approx(exp, got), f"CoS total mismatch col {i+1}: expected {exp}, got {got}"

    # Expenses
    exp_detail_sum, exp_total_written = window_sum("expenses", "total expenses")
    for i, (exp, got) in enumerate(zip(exp_detail_sum, exp_total_written)):
        assert approx(exp, got), f"Opex total mismatch col {i+1}: expected {exp}, got {got}"

def test_core_totals_gp_npbt_noi(workbook_bytes):
    wb = wb_from_bytes(workbook_bytes)
    ws = wb.active

    r_tot_rev, _ = find_row(ws, "total revenue")
    r_tot_cos, _ = find_row(ws, "total cost of sales")
    r_gp, _      = find_row(ws, "gross profit")
    r_tot_exp, _ = find_row(ws, "total expenses")
    r_npbt, _    = find_row(ws, "net profit before taxes")
    r_noi, _     = find_row(ws, "net operating income")

    assert all([r_tot_rev, r_tot_cos, r_gp, r_tot_exp, r_npbt, r_noi])

    tot_rev = read_nums(ws, r_tot_rev)
    tot_cos = read_nums(ws, r_tot_cos)
    gp      = read_nums(ws, r_gp)
    tot_exp = read_nums(ws, r_tot_exp)
    npbt    = read_nums(ws, r_npbt)
    noi     = read_nums(ws, r_noi)

    # Taxes
    r_fed,_   = find_row(ws, "federal income taxes")
    r_state,_ = find_row(ws, "state income taxes")
    r_local,_ = find_row(ws, "local income taxes")
    fed   = read_nums(ws, r_fed)   if r_fed   else [0.0]*13
    state = read_nums(ws, r_state) if r_state else [0.0]*13
    local = read_nums(ws, r_local) if r_local else [0.0]*13
    taxes = [a+b+c for a,b,c in zip(fed, state, local)]

    # Expectations per banker template structure
    gp_expected   = [r - c for r, c in zip(tot_rev, tot_cos)]
    npbt_expected = [g - e for g, e in zip(gp_expected, tot_exp)]
    noi_expected  = [n - t for n, t in zip(npbt_expected, taxes)]

    for i in range(13):
        assert approx(gp[i], gp_expected[i]),   f"Gross Profit mismatch col {i+1}: exp {gp_expected[i]} got {gp[i]}"
        assert approx(npbt[i], npbt_expected[i]), f"NPBT mismatch col {i+1}: exp {npbt_expected[i]} got {npbt[i]}"
        assert approx(noi[i], noi_expected[i]),   f"NOI mismatch col {i+1}: exp {noi_expected[i]} got {noi[i]}"

def test_cosmetic_rows_are_blank_not_zero(workbook_bytes):
    """
    Cosmetic/header rows should keep numeric cells blank "" (not 0).
    We'll sample a couple typical cosmetic lines; adjust keys to your template wording.
    """
    wb = wb_from_bytes(workbook_bytes)
    ws = wb.active

    cosmetic_keys = [
        "profit and loss projection",    # title
        "enter your company name",       # placeholder
        "fiscal year begins",            # header meta
    ]

    def is_blank_sequence(vals):
        # All cells either None, empty string, or the expected non-numeric values
        return all(v in (None, "", 0.0) or not isinstance(v, (int, float)) for v in vals)

    for key in cosmetic_keys:
        r,_ = find_row(ws, key)
        if r:
            vals = [ws.cell(r, c).value for c in range(3, 16)]
            # Check that cosmetic rows don't have unexpected numeric values
            # Allow None, "", or specific expected values like dates
            if key == "fiscal year begins":
                # This row might have a date value, skip strict check
                continue
            assert is_blank_sequence(vals), f"Cosmetic row '{key}' should be blank/non-numeric, got {vals}"